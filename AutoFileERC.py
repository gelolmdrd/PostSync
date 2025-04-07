import tkinter as tk
from tkinter import filedialog, messagebox
from openpyxl import load_workbook
from datetime import datetime
import os

# Main database file path
DATABASE_PATH = "BAC Storage Inventory Database.xlsx"
DATABASE_SHEET = "Sheet1"  # Update if needed

def process_input_file():
    # Let user select the input file
    input_file = filedialog.askopenfilename(title="Select Cabinet File", filetypes=[("Excel Files", "*.xlsx")])
    if not input_file:
        return

    try:
        input_wb = load_workbook(input_file)
        input_ws = input_wb.active

        db_wb = load_workbook(DATABASE_PATH)
        db_ws = db_wb[DATABASE_SHEET]

        count = 0

        for row in input_ws.iter_rows(min_row=2, values_only=True):
            lot_no = row[0]
            title_and_remark = row[1]

            if not lot_no or not title_and_remark:
                continue

            # Split title and remark
            if " - " in title_and_remark:
                title, remark = title_and_remark.split(" - ", 1)
            else:
                title = title_and_remark
                remark = ""

            # Append to database
            next_row = db_ws.max_row + 1
            db_ws.cell(row=next_row, column=1).value = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            db_ws.cell(row=next_row, column=2).value = lot_no
            db_ws.cell(row=next_row, column=3).value = title
            db_ws.cell(row=next_row, column=4).value = remark
            count += 1

        db_wb.save(DATABASE_PATH)
        messagebox.showinfo("Success", f"{count} entries added to the database.")

    except Exception as e:
        messagebox.showerror("Error", str(e))

# UI
root = tk.Tk()
root.title("BAC Excel Importer")
root.geometry("400x150")

tk.Label(root, text="Import Excel entries to BAC Database", font=("Arial", 12)).pack(pady=10)
tk.Button(root, text="Select Input Excel File", command=process_input_file).pack(pady=20)

root.mainloop()
